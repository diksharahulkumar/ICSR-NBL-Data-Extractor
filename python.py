#  ONLY change this line — folder where your xlsx files are kept
INPUT_FOLDER = r"C:\Users\HP\Downloads\(No subject)"

# Output folder is created automatically beside the input folder
OUTPUT_FOLDER_NAME = "NBL_Output"
OUTPUT_FILE_NAME   = "NBL_Summary.xlsx"
import re
import traceback
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell


def get_cell_value(ws, row, col_letter_str):
    """Fetch cell value safely, resolving merged cells to their master cell."""
    col_idx = ord(col_letter_str.upper()) - ord('A') + 1
    cell = ws.cell(row=row, column=col_idx)
    if isinstance(cell, MergedCell):
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= row <= mr.max_row and mr.min_col <= col_idx <= mr.max_col:
                return ws.cell(row=mr.min_row, column=mr.min_col).value
        return None
    return cell.value


def find_field(ws, label_lower, search_col, value_col):
    """Scan every row; when search_col matches the label, return value_col of that row."""
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if get_column_letter(cell.column) == search_col.upper() and cell.value:
                if str(cell.value).strip().lower() == label_lower:
                    return get_cell_value(ws, cell.row, value_col)
    return None


def parse_amount(val):
    """Convert amount to int — handles int, float, or Indian string like '15,00,000'."""
    if val is None:
        return "NOT FOUND"
    if isinstance(val, (int, float)):
        return int(val)
    cleaned = re.sub(r'[,\s]', '', str(val))
    try:
        return int(float(cleaned))
    except ValueError:
        return str(val)


def extract_from_sheet(ws):
    """Extract all required fields from a single worksheet."""
    nbl_no  = find_field(ws, "number & date",               search_col="A", value_col="C")
    pi_name = find_field(ws, "requested by & date:",        search_col="B", value_col="D")
    proj_no = find_field(ws, "project number:",             search_col="B", value_col="D")
    amount  = find_field(ws, "requested negative balance:", search_col="B", value_col="D")
    return {
        "Project No":           str(proj_no).strip()  if proj_no  else "NOT FOUND",
        "Negative Balance No.": str(nbl_no).strip()   if nbl_no   else "NOT FOUND",
        "PI Name":              str(pi_name).strip()  if pi_name  else "NOT FOUND",
        "Amount":               parse_amount(amount),
    }


def extract_from_file(filepath):
    """
    Extract records from ALL sheets in the xlsx file.
    Returns a list of dicts (one per sheet that contains NBL data).
    """
    wb = load_workbook(filepath, data_only=True)
    records = []
    for sheet_name in wb.sheetnames:
        ws  = wb[sheet_name]
        rec = extract_from_sheet(ws)
        # Only include sheet if at least NBL No. or Project No. was found
        if rec["Negative Balance No."] != "NOT FOUND" or rec["Project No"] != "NOT FOUND":
            rec["_sheet"] = sheet_name
            records.append(rec)
    return records


def build_summary_excel(records, output_path):
    """Write a formatted Excel summary from extracted records."""
    wb_out = Workbook()
    ws_out = wb_out.active
    ws_out.title = "NBL Summary"

    thin        = Side(style="thin", color="000000")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill    = PatternFill("solid", fgColor="4472C4")
    hdr_font    = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    yellow_fill = PatternFill("solid", fgColor="FFFF00")
    center_aln  = Alignment(horizontal="center", vertical="center")
    left_aln    = Alignment(horizontal="left",   vertical="center")

    headers = ["Sl No", "Project No", "Negative Balance No.", "PI Name", "Amount"]
    ws_out.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws_out.cell(row=1, column=col_idx)
        cell.fill      = hdr_fill
        cell.font      = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws_out.row_dimensions[1].height = 24

    for rec in records:
        ws_out.append([rec.get(h, "") for h in headers])
        row_idx = ws_out.max_row
        for col_idx, header in enumerate(headers, 1):
            cell = ws_out.cell(row=row_idx, column=col_idx)
            cell.border    = border
            cell.alignment = center_aln if header in ("Sl No", "Negative Balance No.", "Amount") else left_aln
            if header in ("Project No", "Negative Balance No.", "PI Name", "Amount"):
                cell.fill = yellow_fill
            if header == "Amount" and isinstance(cell.value, int):
                cell.number_format = "#,##0"

    for col_idx, width in enumerate([7, 28, 22, 30, 16], 1):
        ws_out.column_dimensions[get_column_letter(col_idx)].width = width
    ws_out.freeze_panes = "A2"
    wb_out.save(output_path)


print(" All functions loaded successfully.")

# Resolve paths
input_folder  = Path(INPUT_FOLDER).resolve()
output_folder = input_folder.parent / OUTPUT_FOLDER_NAME
output_path   = output_folder / OUTPUT_FILE_NAME

# Validate input folder
if not input_folder.exists():
    print(f" ERROR: Folder not found:\n   {input_folder}")
    print("   Please update INPUT_FOLDER in Cell 2 and re-run.")
else:
    # Create output folder automatically
    output_folder.mkdir(parents=True, exist_ok=True)

    # Collect xlsx files:
    #   - Skip ~$ temp files (Excel lock files created when file is open)
    #   - Skip any existing summary output file
    xlsx_files = sorted([
        f for f in input_folder.glob("*.xlsx")
        if not f.name.startswith("~$")          # ← fix: skip Excel temp lock files
        and OUTPUT_FILE_NAME not in f.name
    ])

    if not xlsx_files:
        print(f" ERROR: No .xlsx files found in:\n   {input_folder}")
    else:
        print(f"{'='*62}")
        print(f"  NBL Data Extractor  —  IIT Madras IC&SR")
        print(f"{'='*62}")
        print(f"  Input  folder : {input_folder}")
        print(f"  Output folder : {output_folder}")
        print(f"  Files found   : {len(xlsx_files)}")
        print(f"{'='*62}\n")

        all_records = []
        sl_no = 1
        ok_count, warn_count, fail_count = 0, 0, 0

        for i, filepath in enumerate(xlsx_files, 1):
            print(f"  [{i:>3}/{len(xlsx_files)}]  {filepath.name}")
            try:
                file_records = extract_from_file(filepath)

                if not file_records:
                    print(f"         └─   No NBL data found in any sheet")
                    warn_count += 1
                    continue

                for rec in file_records:
                    rec["Sl No"] = sl_no
                    sl_no += 1
                    missing = [k for k, v in rec.items() if v == "NOT FOUND" and k not in ("Sl No", "_sheet")]
                    sheet_label = f"Sheet: {rec.get('_sheet', '?')}"
                    if missing:
                        print(f"         └─   [{sheet_label}]  WARN: {missing} not found")
                        warn_count += 1
                    else:
                        print(f"         └─   [{sheet_label}]  OK")
                        ok_count += 1
                    all_records.append(rec)

            except Exception as e:
                print(f"         └─  FAILED: {e}")
                traceback.print_exc()
                all_records.append({
                    "Sl No": sl_no, "Project No": "ERROR",
                    "Negative Balance No.": "ERROR",
                    "PI Name": str(e), "Amount": "ERROR",
                })
                sl_no += 1
                fail_count += 1

        build_summary_excel(all_records, output_path)

        print(f"\n{'='*62}")
        print(f"   Done!   {ok_count} OK   {warn_count} warnings   {fail_count} failed")
        print(f"  Total rows in summary : {len(all_records)}")
        print(f"   Saved -> {output_path}")
        print(f"{'='*62}")

import pandas as pd

df = pd.read_excel(output_path)
df["Amount"] = df["Amount"].apply(
    lambda x: f"\u20b9 {int(x):,}" if isinstance(x, (int, float)) and str(x).replace('.','').isdigit() else x
)

print(f"Total records: {len(df)}\n")
df.style.set_properties(**{"text-align": "left"}).set_table_styles(
    [{"selector": "th", "props": [("text-align", "center"),
                                   ("background-color", "#4472C4"),
                                   ("color", "white")]}]
)